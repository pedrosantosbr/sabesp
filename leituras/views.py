from django.shortcuts import render, redirect
from leituras.forms import UploadLeituraForm
from leituras.models import Folha, Leitura, Relatorio, RelatorioEvento, Condominio
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.conf import settings
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
from datetime import datetime
from rest_framework.exceptions import ValidationError


@login_required(login_url=settings.LOGIN_URL)
@transaction.atomic
def add_leitura(request):
    if request.method == "POST":
        form = UploadLeituraForm(request.POST, request.FILES)
        if form.is_valid():
            wb = load_workbook(request.FILES["file"])
            sheet = wb.active

            condominio_nome = form.cleaned_data["condominio_nome"]
            if condominio_nome and condominio_nome != "":
                condominio_nome = condominio_nome.lower().strip()
            else:
                raise Exception("Nome do condomínio não pode ser vazio")

            cm, _ = Condominio.objects.get_or_create(nome=condominio_nome)
            rl = Relatorio.objects.create(condominio=cm)
            fl = Folha.objects.create(relatorio=rl, arquivo=request.FILES["file"])

            for i in range(5, sheet.max_row + 1):
                row = sheet[i]

                try:
                    date_string = str(row[5].value).split(" ")[0]
                    time_string = str(row[6].value)
                    print(time_string)
                except Exception:
                    raise ValidationError(detail="Erro ao ler data e hora da planilha")

                try:
                    datetime_string = date_string + " " + time_string
                    formatted_datetime = datetime.strptime(
                        datetime_string, "%Y-%m-%d %H:%M:%S"
                    )
                except Exception:
                    raise ValidationError(detail="Erro ao formatar data e hora")

                Leitura.objects.create(
                    folha=fl,
                    rgi_principal=row[0].value,
                    rgi_autonoma=row[0].value,
                    status_valvula="",
                    data_leitura=formatted_datetime,
                    leitura=row[4].value,
                    tipo=row[3].value.split("-")[1],
                    imovel=row[2].value,
                    hidrometro=row[1].value,
                )

            RelatorioEvento.objects.create(
                funcionario=request.user,
                relatorio=rl,
                event_type=RelatorioEvento.EventTypes.CRIADO,
            )
            return redirect("dashboard")
    else:
        form = UploadLeituraForm()
    return render(request, "leituras/add.html", {"form": form})


@login_required(login_url=settings.LOGIN_URL)
@transaction.atomic
def edit_leitura(request, id):
    relatorio = get_object_or_404(Relatorio, pk=id)

    if request.method == "POST":
        form = UploadLeituraForm(request.POST, request.FILES)
        if form.is_valid():
            relatorio.folhas.all().delete()

            fl = Folha.objects.create(
                relatorio=relatorio, arquivo=request.FILES["file"]
            )
            wb = load_workbook(request.FILES["file"])
            sheet = wb.active

            for row in sheet.iter_rows():
                Leitura.objects.create(
                    folha=fl,
                    rgi_principal=row[0].value,
                    rgi_autonoma=row[1].value,
                    status_valvula="Aberta",
                    data_leitura=row[5].value,
                    leitura=row[4].value,
                    fria_quente=row[3].value.split("-")[1],
                    imovel=row[2].value,
                    hidrometro="AZ000000",
                )

            RelatorioEvento.objects.create(
                funcionario=request.user,
                relatorio=relatorio,
                event_type=RelatorioEvento.EventTypes.ATUALIZADO,
            )
            messages.success(request, "Leitura atualizada com sucesso!")
            return redirect("dashboard")
    else:
        form = UploadLeituraForm()

    return render(request, "leituras/edit.html", {"form": form, "relatorio": relatorio})


@login_required(login_url=settings.LOGIN_URL)
@transaction.atomic
def delete_relatorio(request, id):
    relatorio = get_object_or_404(Relatorio, pk=id)
    if request.method == "POST":
        relatorio.folhas.all().delete()
        relatorio.deleted_at = timezone.now()
        relatorio.save()

        messages.success(request, "Leitura deletada com sucesso!")

        return redirect("dashboard")
    return render(request, "leituras/delete.html", {"relatorio": relatorio})
